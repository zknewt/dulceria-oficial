from io import BytesIO
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def queryset_to_excel(filename, columns, rows_iter):

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Estilos simples
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44546A")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezados
    ws.append([h for (h, _) in columns])
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Filas
    rownum = 2
    for obj in rows_iter:
        row = []
        for _, extractor in columns:
            try:
                val = extractor(obj)
            except Exception:
                val = ""
            # Hint de formatos: fecha, Decimal, bool → Excel-friendly
            if isinstance(val, datetime):
                # openpyxl detecta datetime y Excel lo formatea como fecha
                pass
            elif isinstance(val, Decimal):
                val = float(val)
            elif val is None:
                val = ""
            row.append(val)
        ws.append(row)
        # bordes
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=rownum, column=col_idx).border = border
        rownum += 1

    # Auto ancho de columnas
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Guardar a memoria
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)

    final_name = f"{filename}.xlsx"
    return mem.read(), final_name
